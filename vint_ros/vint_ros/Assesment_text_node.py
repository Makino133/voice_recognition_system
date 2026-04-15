
import os
import requests
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import yaml
import re
from google import genai



#pkg_path = get_package_share_directory('vint_ros')
pkg_path="/home/sip-mobility/docker_projects/nav2_ws/src/voice_recognition_system/vint_ros/"

DB_PATH = os.path.join(pkg_path,'others/data_base_100.yaml')
PROMPT_PATH = os.path.join(pkg_path,'others/prompts.yaml')
API_PATH =  os.path.join(pkg_path,'conf/conf.yaml')

with open(DB_PATH, "r") as f:
    db_handlr= yaml.safe_load(f)["db_full"]

with open(PROMPT_PATH, "r") as f:
    prompt_handlr= yaml.safe_load(f)

with open(API_PATH, "r") as f:
    conf_handlr= yaml.safe_load(f)
    API_KEY = conf_handlr["key"]
    API_KEY2 = conf_handlr["key2"]
    API_KEY_OR = conf_handlr["key_or"]




# voice_recognition
class VoiceRecognition():
    def __init__(self):
        print("===================================================")
        self.prompts=prompt_handlr
        self.db=db_handlr
        self.cases = list(self.db.keys())
        self.cs_str=0 # Case to start from
        self.i_str=0  # Index to start from
        self.keys=[]

        if API_KEY:
            self.source = "Genai"
            self.keys.append(API_KEY)
            if API_KEY2:
                self.keys.append(API_KEY2)
        elif API_KEY_OR:
            self.source="Open Router"
        else:
            raise ValueError("NO API key")
        
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def run(self):     
        TF = "origin"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"prompt_test_{timestamp}.xlsx"
        header = ["Answer Label", "Request", "LLM Output", "Judged Label", "True/False", "Latency"]

        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(filename)

        for lab , exs in self.db.items():
            wb = load_workbook(filename)
            ws = wb.active

            if self.cases.index(lab)<self.cs_str:
                continue

            for i, text in enumerate(exs):
                
                if self.cases.index(lab)<(self.cs_str+1) and i < self.i_str:
                    continue
                print(f"Label: {lab}, Example {i}")
                start= datetime.now()
                out_LLM = self.LLM(text,"ppose_select_both_examples_light_IDK_exs_cons_claude_extra")
                lat= (datetime.now()-start).total_seconds()
                judged_label= self.result_edge(out_LLM)
                TF= int(judged_label == lab)

                ws.append([lab, text, out_LLM, judged_label, TF , lat])

                wb.save(filename)

                print("------------------------")
                time.sleep(3)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def LLM(self, text , prompt_key):

            base_prompt= self.prompts[prompt_key]
            prompt = base_prompt + text

            if self.source=="Open Router":
                API_URL = "https://openrouter.ai/api/v1/chat/completions"
                headers = {"Authorization": f"Bearer {API_KEY_OR}","Content-Type": "application/json"}
                data = {"model": "google/gemma-3n-7b-it:free","messages": [{"role": "user", "content": prompt}]}
                
                result = requests.post(API_URL, json=data, headers=headers)
                print(result)

                try:
                    output = result["choices"][0]["message"]["content"]
                except KeyError:
                    print("KeyError in result, full content:", result)
                except TypeError:
                    print("PrintError in result, full content:", result)


            else:
                key=self.keys.pop(0)
                client = genai.Client(api_key=key)
                response = client.models.generate_content(model="gemma-3n-e4b-it", contents=prompt)
                #response = self.call_with_retry(client.models.generate_content,model="gemma-3-4b", contents=prompt)
                output = response.text
                self.keys.append(key)
            

                print("Assistant:", output)
            return output
 

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def result_edge(self, LLM_out):

        self.out_edge = None
   
        LLM_lower = LLM_out.split(".")[-2].casefold()  #Picking the last sentence of the LLM response

        for labels in self.cases:
            parts = [p.strip() for p in labels.split(" / ")]
            for part in parts:
                pattern = rf"\b{re.escape(part.casefold())}\b"
                if re.search(pattern, LLM_lower):
                    

                    return part

        if self.out_edge is None:
            print(f"""No matching edge found in LLM output: "{LLM_lower}" """)
            #self.get_logger().warn(f"""No matching edge found in LLM output: "{LLM_lower}" """)
        else:
            print(f"(Matching Edge: {self.out_edge}")
            #self.get_logger().info(f"(Matching Edge: {self.out_edge}")
        return None
    
    def call_with_retry(func, *args, max_retries=5, **kwargs):
        """Call a Gemini API function with automatic retry on 429 errors."""
        for attempt in range(max_retries):
            try:
                return func(*args, **kwargs)
            
            except genai.errors.ClientError as e:
                if e.code != 429:
                    raise  # re-raise if it's not a quota error

                # Extract retry delay from error message if available
                match = re.search(r'retry in (\d+(?:\.\d+)?)s', str(e))
                wait = float(match.group(1)) if match else 60

                print(f"[Attempt {attempt+1}/{max_retries}] Rate limited. Waiting {wait:.1f}s...")
                time.sleep(wait + 1)  # +1s buffer

        raise RuntimeError(f"Failed after {max_retries} retries.")

##############################################################
def main():
    voice = VoiceRecognition()
    voice.run()

##############################################################
if __name__ == "__main__":
    main()
