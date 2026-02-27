#!/usr/bin/env python3

import json
import vosk
import queue
import threading
import rclpy
import time
import numpy as np
import pyttsx3
import math
import random
import time
import re
from std_msgs.msg import Float32MultiArray
from std_msgs.msg import Empty
from rclpy.node import Node
from rclpy.parameter import Parameter as RclpyParameter
from rclpy.executors import MultiThreadedExecutor
from rcl_interfaces.msg import Parameter as MsgParameter, ParameterType, ParameterValue
from rcl_interfaces.srv import SetParameters
from ament_index_python.packages import get_package_share_directory
from geometry_msgs.msg import TransformStamped, PoseStamped
from visualization_msgs.msg import Marker
from tf2_ros import TransformBroadcaster
from builtin_interfaces.msg import Duration
from geometry_msgs.msg import Point
from std_msgs.msg import String
from std_msgs.msg import Int32
from openpyxl import Workbook, load_workbook
from datetime import datetime
import ast

#----------------------------------------------------
class Assessment(Node):
    def __init__(self):
        super().__init__('Assessment')
        self.get_logger().info("==================================================\n")
        self.goal_num = 1
        self.edge_x = 0.0
        self.edge_y = 0.0
        self.vis = 0.0
        self.command = None
        self.LLM_out = None
        self.edge_list = None
        self.out_edge_num = None
        self.prediction_requested = False
        self.state = {"table": {"pos": None, "quat": None, "size": None},}
        self.target_edge= None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"prompt_test_{timestamp}.xlsx"
        header = ["Target Label", "Voice Command" , "LLM Selection", "LLM response", "Correct predictions?"]
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(self.filename)

        #---- publish ----#
        self.target_point_pub = self.create_publisher(Marker, '/target_point', 10)
        self.arrow_marker = self.create_timer(1.0, self.publish_arrow_marker)    
        self.LLM_out_line_pub = self.create_publisher(Marker, '/LLM_out_line', 10)
        self.LLM_marker = self.create_timer(0.1, self.publish_arrow_marker)    
        self.map_update_pub = self.create_publisher(Float32MultiArray, '/map_update', 10)
        self.cposes_pub = self.create_publisher(Marker, "/select_poses", 10) 

        # ---- subscribe ----
        self.sub_table = self.create_subscription(Marker,
                "/table_marker",
                lambda msg: self.marker_cb(msg, "table"),
                10
                )

        self.command_sub = self.create_subscription(
                String,
                "/command",
                self.command_cb,
                10
                )
        self.LLM_out_sub = self.create_subscription(
                String,
                "/LLM_out",
                self.LLM_out_cb,
                10
                )
        self.edge_sub = self.create_subscription(
                String,
                "/edge_list",
                self.edge_cb,
                10
                )
        self.prediction_sub = self.create_subscription(
                Empty,
                '/prediction',
                self.prediction_cb,
                10
                )
                
#=============== "CB" =================================================
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~               
    def marker_cb(self, msg: Marker, object_name: str):
        if object_name not in self.state:
            return

        p = msg.pose.position
        q = msg.pose.orientation
        s = msg.scale

        curr = (
            p.x, p.y, p.z,
            q.x, q.y, q.z, q.w,
            s.x, s.y, s.z
        )
        
        self.state[object_name]["pos"]  = (p.x, p.y, p.z)
        self.state[object_name]["quat"] = (q.x, q.y, q.z, q.w)
        self.state[object_name]["size"] = (s.x, s.y, s.z)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def prediction_cb(self, msg):

        self.get_logger().info("prediction received")

        self.prediction_requested = True
        time.sleep(3.0)
        self.try_process()

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def command_cb(self, msg: String):
        self.command = msg.data.strip()
        if self.command == None:
            self.get_logger().info("command is None")
            return
        self.get_logger().info(f"command received: {self.command}")

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def LLM_out_cb(self, msg: String):
        lab_reps = msg.data.split(" // ")
        self.LLM_lab = lab_reps[0]
        self.LLM_out = lab_reps[1]
        self.get_logger().info("LLM out received")

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def edge_cb(self, msg: String):
        edge_text = msg.data.strip()

        edges_nm_ps = [s.strip() for s in edge_text.splitlines() if s.strip()]
        
        self.edge_list = {}

        if len(edges_nm_ps) != 4:
            self.get_logger().warn(f"Unexpected /edge_list format: {edges_nm_ps}")
            return
        
        for edge_nm_ps in edges_nm_ps:
            edge_name = edge_nm_ps.split("//")[0]
            edge_ps = (ast.literal_eval(edge_nm_ps.split("//")[-1]))
            self.edge_list.update({edge_name : edge_ps})


#======================================================================

    def try_process(self):

        if not self.prediction_requested:
            return
        if self.LLM_out is None:
            self.get_logger().info("LLM_out not received yet")
            return
        if self.edge_list is None:
            self.get_logger().info("edge_list not received yet")
            return
        if self.command is None:
            self.get_logger().info("command not received yet")
            return

        self.get_logger().info("")
        self.get_logger().info("--------------- Result ----------------")

        # command出力

        # 設定ゴール番号 → ゴールエッジラベル
        self.get_logger().info(f"Goal Edge: {self.target_edge}")

        self.get_logger().info(f"Voice Command: {self.command}")


        self.get_logger().info(f"LLM selection: {self.LLM_lab}")


        self.get_logger().info(f"LLM response: {self.LLM_out}")


        T_or_F = self.LLM_lab.casefold() in self.target_edge.casefold()

        self.get_logger().info(f"Result: {T_or_F}")
        
        # 結果の保存
        self.save_data(self.target_edge , self.command, self.LLM_lab, self.LLM_out, T_or_F)
        self.get_logger().info("(Save result)")
        self.get_logger().info("---------------------------------------")
        # LLM_lineの表示
        self.vis = 1.0
        #time.sleep(5)
        self.vis = 0.0
        self.get_logger().info("")
        
        #~~~~~~~~~~~~~~次のループ開始~~~~~~~~~~~~~~~~~~~~#
        values = Float32MultiArray()
        self.goal_num = random.randint(0, 3)
        while True:
            self.robot_x = round(random.uniform(-3, 3), 1)
            self.robot_y = round(random.uniform(-3, 3), 1)
            self.robot_yaw = round(random.uniform(0, 359), 1)
            self.table_size_x = round(random.uniform(0.5, 1.2), 1)
            self.table_size_y = round(random.uniform(1.0, 1.5), 1)
            self.table_size_h = round(random.uniform(0.5, 0.7), 1)
            dx=self.robot_x
            dy=self.robot_y
            t_id=math.atan2(dy,dx)*180/3.14
            self.robot_yaw = round(random.uniform(t_id-10, t_id+10), 1)

            if (abs(self.robot_x) > 1.0) and (abs(self.robot_y) > 1.0):
                break
        self.get_logger().info("------ Robot position is updated ------")
        self.get_logger().info(f"robot_x: {self.robot_x}")
        self.get_logger().info(f"robot_y: {self.robot_y}")
        self.get_logger().info("---------------------------------------\n")
        values.data = [self.robot_x, self.robot_y, self.robot_yaw, 
                self.table_size_x, self.table_size_y, self.table_size_h, 1.0]
        self.map_update_pub.publish(values)

        msg = Marker()
        msg.action = Marker.DELETEALL
        self.cposes_pub .publish(msg)

        # ===== リセット =====
        self.prediction_requested = False
        self.LLM_out = None


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def save_data(self, goal_edge, command, LLM_edge , LLM_out, TF):
        wb = load_workbook(self.filename)
        ws = wb.active
        ws.append([ goal_edge[1], command, LLM_edge, LLM_out, TF])
        wb.save(self.filename)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def publish_arrow_marker(self):
        # ----- arrow ------#
        now = self.get_clock().now().to_msg()
        if self.edge_list is None:
            return

        arrow = Marker()
        arrow.header.frame_id = "map"
        arrow.header.stamp = now
        arrow.ns = "target_point"
        arrow.id = 1
        arrow.type = Marker.ARROW
        arrow.action = Marker.ADD


        self.target_edge = list(self.edge_list.keys())[self.goal_num]
        posex , posey ,yaw = self.edge_list[self.target_edge]

        # ✅ position from poses[name]
        arrow.pose.position.x = posex
        arrow.pose.position.y = posey
        arrow.pose.position.z = 0.0

        # ✅ orientation from yaw
        
        q=self.yaw_to_quaternion(yaw)
        arrow.pose.orientation.x = q[0]
        arrow.pose.orientation.y = q[1] 
        arrow.pose.orientation.z = q[2] 
        arrow.pose.orientation.w = q[3] 

        # Arrow size
        arrow.scale.x = 0.6   # length
        arrow.scale.y = 0.1   # width
        arrow.scale.z = 0.1   # height

        # 矢印の太さ
        arrow.scale.x = 0.6  # シャフト径
        arrow.scale.y = 0.1   # ヘッド径
        arrow.scale.z = 0.1   # ヘッド長
        arrow.color.r = 1.0
        arrow.color.g = 0.0
        arrow.color.b = 0.0
        arrow.color.a = 1.0

        self.target_point_pub.publish(arrow)
        # ------------------#

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def yaw_to_quaternion(eslf, yaw: float) -> list:
        """Convert yaw angle (rad) to quaternion."""
        qx = 0.0
        qy = 0.0
        qz = math.sin(yaw / 2.0)
        qw = math.cos(yaw / 2.0)
        q=[qx,qy,qz,qw]
        return q



def main():
    rclpy.init()
    assessment = Assessment()
    rclpy.spin(assessment)

if __name__=="__main__":
    main()
