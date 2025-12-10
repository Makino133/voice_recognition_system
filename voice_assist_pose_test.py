import sys
import os
import json
import vosk
import sounddevice as sd
import queue
import threading
import rclpy
import requests
import time
import numpy as np
import pyttsx3
import math
import random
import pandas as pd
from rclpy.node import Node
from rclpy.parameter import Parameter as RclpyParameter
from rclpy.executors import MultiThreadedExecutor
from rcl_interfaces.msg import Parameter as MsgParameter, ParameterType, ParameterValue
from rcl_interfaces.srv import SetParameters
from vosk import Model, KaldiRecognizer , GpuInit
from ament_index_python.packages import get_package_share_directory
from geometry_msgs.msg import TransformStamped, PoseStamped
from visualization_msgs.msg import Marker
from tf2_ros import TransformBroadcaster
from builtin_interfaces.msg import Duration
from geometry_msgs.msg import Point
from openpyxl import Workbook, load_workbook
from datetime import datetime

GpuInit()

#--------------------------------------------------
#MODEL_PATH = "/home/orin/voice_recognition_system/vosk-model-lm/vosk-model-en-us-0.22"
#model = Model(MODEL_PATH)

API_KEY = "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"
API_URL = "https://openrouter.ai/api/v1/chat/completions"

engine = pyttsx3.init()
engine.setProperty("rate", 150)
engine.setProperty("volume", 1.0)

headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
        }

q = queue.Queue()
#---------------------------------------------------
robot_x = 1.5
robot_y = 4.5
robot_yaw_deg = 270.0
robot_yaw_rad = math.radians(robot_yaw_deg)
robot_name = "robot"
Robot = np.array([robot_x, robot_y, robot_yaw_deg, robot_yaw_rad, robot_name]) 
robot_size_x = 0.5
robot_size_y = 0.5
robot_size_z = 1.0

table_x = 2.0
table_y = 3.0
table_yaw_deg = 0.0
table_yaw_rad = math.radians(table_yaw_deg)
table_name = "table"
Table = np.array([table_x, table_y, table_yaw_deg, table_yaw_rad, table_name]) 
table_size_x = 1.5
table_size_y = 0.9
table_size_z = 0.75

THETA_A = 15.0

table_hsize_x = table_size_x / 2
table_hsize_y = table_size_y / 2
corners = [
        ((table_x + table_hsize_x), (table_y + table_hsize_y), table_size_z),
        ((table_x - table_hsize_x), (table_y + table_hsize_y), table_size_z),
        ((table_x - table_hsize_x), (table_y - table_hsize_y), table_size_z),
        ((table_x + table_hsize_x), (table_y - table_hsize_y), table_size_z)
        ]
corner_ijkl = np.zeros((4,5))
#----------------------------------------------------
edge_flag = 0
req_edge = 0
target = "None"

#####################################

labels = ["Near Edge", "Right Edge", "Left Edge", "Far Edge", "Near Right Edge", "Far Right Edge", "Far Left Edge", "Near Left Edge"]

count_false = np.zeros(8)

req_Near = ["I’d like to move a bit closer to the table", 
	"I want to move toward the near side", 
	"I’d like to move forward from here", 
	"I want the spot where I can reach easily", 
	"The nearer position is easier for working", 
	"I’d like to sit toward the front", 
	"I’d like to bring my body closer to the table", 
	"The near edge is easier for me to use", 
	"I feel more comfortable around here", 
	"This side feels more natural"
        ]
        
req_Far = ["I want to be a bit farther from the table", 
	"I’d like to sit toward the far side", 
	"I’d like to maintain some distance", 
	"I’d like to move around to the far side", 
	"I’d like to move back a little", 
	"The far position feels calmer", 
	"I’d like the side that gives me a wide view", 
	"I want to work from the far side", 
	"Being a little away feels easier", 
	"The far area feels more spacious"
        ]
        
req_Right = ["I want to move to the right side", 
	"I’d like to move around to the right", 
	"I want to go to the right edge", 
	"I’d like the side where I can use my right hand easily", 
	"I’d like to shift slightly to the right from here", 
	"I want to sit facing to the right", 
	"The right side surface is easier to work with", 
	"My posture feels more stable when shifted right", 
	"That direction feels natural", 
	"The right direction feels right"
        ]
        
req_Left = ["I want to move to the left side", 
	"I’d like to move around to the left", 
	"I want to go to the left edge", 
	"I’d like the side where I can use my left hand easily", 
	"I’d like to shift a little to the left from here", 
	"I want to sit facing to the left",
	"The left side surface is easier to work with", 
	"My posture feels natural on the left side", 
	"The left direction feels calming", 
	"This orientation feels right"
        ]    

req_NearRight = ["I want to move to the near right side",
	"I’d like the near right spot where I can rest my hand", 
	"I want to sit closer on the right side", 
	"I’d like to shift slightly to the right from here", 
	"The near right edge is easier for me to reach", 
	"The near right area is easier for working", 
	"I want to come around to the near right side", 
	"I’d like the spot where my right hand moves easily", 
	"I can naturally face that side", 
	"This general area feels comfortable"
        ]
        
req_FarRight = ["I want to move to the far right side", 
	"I’d like to sit at the far right", 
	"The far right side feels more comfortable", 
	"I want to work at the far right edge of the table", 
	"I’d like to shift back and orient right", 
	"The far right provides a better view", 
	"I can use more space if I’m on the far right", 
	"I want to keep some distance while staying to the right", 
	"I feel more stable in that direction", 
	"Being toward the far side feels calming"
        ]
        
req_FarLeft = ["I want to move to the far left side", 
	"I’d like to sit at the far left", 
	"I want to work at the far left edge", 
	"I can orient my body more easily at the far left", 
	"I can move my arm comfortably there", 
	"I feel more stable at the far left side", 
	"The far left provides a usable arrangement", 
	"I want a spot where I can settle my gaze easily", 
	"That side feels natural", 
	"That area feels right"
        ]

req_NearLeft = ["I want the near left spot", 
	"I prefer the near left side where I can rest my hand", 
	"I’d like to shift slightly left from here", 
	"The near left area is easy to move in", 
	"I want to come closer and orient left", 
	"I want the position where my left hand works naturally", 
	"The near left area feels comfortable", 
	"I’d like a spot where I can sit left without moving far", 
	"This general area feels easy to be in", 
	"This spot feels right"
        ]       
       
req_false_case1 = ["This orientation feels right",
        "This side feels more natural"
        ]
label_false_case1 = [2, 0]

req_false_case2 = ["I’d like to shift back and orient right",
        "This general area feels comfortable",
        "I can naturally face that side",
        "This spot feels right",
        "I want to sit closer on the right side",
        "I feel more stable in that direction",
        "That area feels right"
        ]
label_false_case2 = [5, 4, 4, 7, 4, 5, 6]

############################################################################

# voice_recognition
class VoiceRecognition():
    def __init__(self):
        print("===================================================")
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def run(self):
        a = 100
        true = 0
        false = 0 
        idk = 0
        fc_8 = 0
        fc_9 = 0
        TF = "origin"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"prompt_test_{timestamp}.xlsx"
        header = ["Scenario", "Answer Label", "Request", "LLM Output", "Judged Label", "True/False"]

        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(filename)

        for i in range(a):
            wb = load_workbook(filename)
            ws = wb.active
            print(i+1)
            scen_num = random.randint(0, 7)
            req_number = random.randint(0, 9)
            """
            if i % 2 == 0:
                scen_num = 8
            elif i % 2 == 1:
                scen_num = 9
            """

            if scen_num == 0:
                text = req_Near[req_number]
            elif scen_num == 1:
                text = req_Right[req_number]
            elif scen_num == 2:
                text = req_Left[req_number]
            elif scen_num == 3:
                text = req_Far[req_number]
            elif scen_num == 4:
                text = req_NearRight[req_number]
            elif scen_num == 5:
                text = req_FarRight[req_number]
            elif scen_num == 6:
                text = req_FarLeft[req_number]
            elif scen_num == 7:
                text = req_NearLeft[req_number]
            """
            elif scen_num == 8:
                fc_8 += 1
                req_number = fc_8 % 2
                #req_number = random.randint(0, 1)
                scen_num = label_false_case1[req_number]
                text = req_false_case1[req_number]
            elif scen_num == 9:
                fc_9 += 1
                req_number = fc_9 % 7
                #req_number = random.randint(0, 6)
                scen_num = label_false_case2[req_number]
                text = req_false_case2[req_number]
            """

            answer = scen_num
            if scen_num < 4:
                scenario = "case1"
                output = self.LLM_edge_case1(text)
                print("output succses")
            elif scen_num > 3:
                scenario = "case2"
                output = self.LLM_edge_case2(text)
                print("output succses")

            edge_num = self.result_edge(output, labels)
            if edge_num == scen_num:
                print(text)
                print(labels[scen_num])
                true += 1
                print("TRUE")
                TF = "True"
                judged_label = labels[edge_num]
            elif edge_num == 99:
                print(text)
                print(labels[scen_num])
                idk += 1
                print("IDK")
                judged_label = "IDK"
                TF = "IDK"
            if not edge_num == scen_num and not TF == "IDK":
                false += 1
                count_false[scen_num] += 1
                print(text)
                print(labels[scen_num])
                print("FALSE")
                TF = "False"
                judged_label = labels[edge_num]

            accuracy = (true / (i+1)) * 100
            print(accuracy, "%")
            print("False count list: ", count_false)
            print("IDK count: ", idk)
            ws.append([scenario, labels[scen_num], text, output, judged_label, TF])

            wb.save(filename)

            print("------------------------")
            time.sleep(10)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def LLM_edge_case1(self, text):
        base_prompt = """
        You are an assistant AI mounted on a wheelchair.

        You are currently near a rectangular table, and there are four possible positions around it:
            - Near Edge – the side of the table closest to the wheelchair
            - Far Edge – the side farthest from the wheelchair
            - Right Edge – the right side of the table relative to the wheelchair’s forward direction
            - Left Edge – the left side of the table relative to the wheelchair’s forward direction

        When the user makes a request, determine the most suitable position only if you can identify it with complete certainty.

        Certainty rules:
        You must be certain only when the request includes:
            - Explicit directional information (e.g., "right", "left", "closest", "far", "back", "front")
            - Unambiguous spatial relationships that match exactly one of the four positions

        If two or more interpretations are possible, or if the information is incomplete, respond “I don't know.”
        Do not infer or guess based on assumptions or probability.

        Output rules:
        Your response must:
            1. Be one or two polite English sentences as if speaking to the user.
            2. Include exactly one of the following labels: “Near Edge”, “Far Edge”, “Right Edge”, “Left Edge”, or “I don't know”.
            3. Never include more than one label.
            4. If uncertain, output exactly 'I don't know.'

        User request:
        """
        prompt = base_prompt + text
        #print(prompt)
        data = {
                "model": "google/gemma-3n-e4b-it:free",
                "messages": [{"role": "user", "content": prompt}]
                }
        response = requests.post(API_URL, json=data, headers=headers)
        print(response)
        result = response.json()
        output = result["choices"][0]["message"]["content"]
        print("Assistant:", output)
        return output

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def LLM_edge_case2(self, text):
        base_prompt = """        
        You are an assistant AI mounted on a wheelchair.

        You are currently near a rectangular table, and there are four possible positions around it:r
                - Near Right Edge – The table edge extending to the right from the corner directly in front of you. It is the closest right-side edge from your perspective.
                - Far Right Edge – The edge opposite the Near Left Edge, located on the far right side of the table. It is the far, right-back edge from your perspective.
                - Far Left Edge – The edge opposite the Near Right Edge, located on the far left side of the table. It is the far, left-back edge from your perspective.
                - Near Left Edge – The table edge extending to the left from the corner directly in front of you. It is the closest left-side edge from your perspective.
        When the user makes a request, determine the most suitable position only if you can identify it with complete certainty.

        Certainty rules:
        You must be certain only when the request includes:
            - Explicit directional information (e.g., "right", "left", "closest", "far", "back", "front")
            - Unambiguous spatial relationships that match exactly one of the four positions

        If two or more interpretations are possible, or if the information is incomplete, respond “I don't know.”
        Do not infer or guess based on assumptions or probability.

        Output rules:
        Your response must:
            1. Be one or two polite English sentences as if speaking to the user.
            2. Include exactly one of the following labels: “Near Right Edge”, “Far Right Edge”, “Near Left Edge”, “Far Left Edge”, or “I don't know”.
            3. Never include more than one label.
            4. If uncertain, output exactly 'I don't know.'

        User request:

        """
        prompt = base_prompt + text
        #print(prompt)
        data = {
                "model": "google/gemma-3n-e4b-it:free",
                "messages": [{"role": "user", "content": prompt}]
                }
        response = requests.post(API_URL, json=data, headers=headers)
        print(response)
        result = response.json()
        output = result["choices"][0]["message"]["content"]
        print("Assistant:", output)
        return output

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def result_edge(self, output, edges):
        edge_number = 0
        for m in range(8):
            if edges[m] in output:
                edge_number = m
                print(edges[m])
                continue
            elif "I don't know" in output:
                edge_number = 99
                print("could not decide edge")
                continue
        print(edge_number)
        return edge_number

##############################################################
def main():
    voice = VoiceRecognition()
    voice.run()

##############################################################
if __name__ == "__main__":
    main()
